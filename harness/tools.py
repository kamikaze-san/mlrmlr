"""Custom tools exposed inside the RLM's REPL. Plain functions -- rlm injects
these into REPL globals via custom_tools={name: fn}."""
import csv
import os
import re
import sqlite3
from pathlib import Path

import openpyxl
from dateutil import parser as dateparser
from openpyxl.utils.cell import range_boundaries

HERE = Path(__file__).resolve().parent
# Dataset location is independent of where this harness folder sits -- set
# DATASET_ROOT to wherever you cloned BITS-Hackathon-Dataset on this machine.
# The fallback only makes sense on the machine this was first written on.
ROOT = Path(os.environ.get("DATASET_ROOT", r"C:\Users\NewGr\Downloads\BITS-Hackathon-Dataset"))
INDEX_CSV = ROOT / "document_index.csv"
DOCS = ROOT / "documents"
CACHE = HERE / "text_cache"
DB_PATH = HERE / "estate.db"

_INDEX = list(csv.DictReader(open(INDEX_CSV, encoding="utf-8")))
_BY_ID = {r["doc_id"]: r for r in _INDEX}
_SUM_RE = re.compile(r"^=SUM\(\s*([A-Za-z]+\d+):([A-Za-z]+\d+)\s*\)$", re.IGNORECASE)


def list_documents(doc_type: str | None = None) -> list[dict]:
    """List documents in the corpus, optionally filtered by doc_type
    (e.g. 'completion_certificate', 'reference_letter', 'personnel_certificate',
    'cv', 'past_performance_portfolio', 'tender_dossier', 'compliance_matrix', ...).
    Returns list of {doc_id, doc_type, filename}."""
    rows = _INDEX if doc_type is None else [r for r in _INDEX if r["doc_type"] == doc_type]
    return [{"doc_id": r["doc_id"], "doc_type": r["doc_type"], "filename": r["filename"]} for r in rows]


def list_doc_types() -> list[str]:
    """Return the distinct document types present in the corpus."""
    return sorted({r["doc_type"] for r in _INDEX})


def read_document(doc_id: str) -> str:
    """Return the full extracted text of one document by its doc_id (e.g. 'DOC-CC-012')."""
    path = CACHE / f"{doc_id}.txt"
    if not path.exists():
        return f"ERROR: no such doc_id '{doc_id}'"
    return path.read_text(encoding="utf-8")


def grep_documents(pattern: str, doc_type: str | None = None, case_sensitive: bool = False) -> list[dict]:
    """Search every document's extracted text for a regex pattern.
    Optionally restrict to one doc_type. Returns a list of
    {doc_id, doc_type, filename, matching_lines: [str]} for every document with
    at least one match -- does not truncate the document list, only the lines shown."""
    flags = 0 if case_sensitive else re.IGNORECASE
    try:
        rx = re.compile(pattern, flags)
    except re.error as e:
        return [{"error": f"bad regex: {e}"}]

    rows = _INDEX if doc_type is None else [r for r in _INDEX if r["doc_type"] == doc_type]
    results = []
    for r in rows:
        path = CACHE / f"{r['doc_id']}.txt"
        if not path.exists():
            continue
        text = path.read_text(encoding="utf-8")
        matches = [line.strip() for line in text.splitlines() if rx.search(line)]
        if matches:
            results.append({
                "doc_id": r["doc_id"],
                "doc_type": r["doc_type"],
                "filename": r["filename"],
                "matching_lines": matches[:15],
            })
    return results


def read_workbook_table(doc_id: str) -> dict:
    """For an xlsx-derived document, read the ORIGINAL workbook directly
    (not the flat tab-separated text cache) and return
    {sheet_name: [{column_name: value, ...}, ...]}, using each sheet's first
    non-empty row as headers. This gives real column-name access instead of
    guessing tab-separated positions in read_document's text -- pass a
    sheet's rows to pandas.DataFrame(rows) for filter/groupby/sum, or just
    use plain Python dict/list comprehensions. Recomputes blank SUM()
    formula cells the same way the text cache extraction does (see
    preprocess.py) so totals are real, not None. Only works for doc_ids
    whose source file is .xlsx -- returns {"error": ...} for anything else."""
    row = _BY_ID.get(doc_id)
    if row is None:
        return {"error": f"no such doc_id '{doc_id}'"}
    filename = row["filename"]
    if not filename.lower().endswith(".xlsx"):
        return {"error": f"'{doc_id}' is not an xlsx document (doc_type={row['doc_type']})"}
    path = DOCS / filename

    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    out = {}
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]
        headers = None
        rows_out = []
        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            values = []
            for cell_v, cell_f in zip(row_v, row_f):
                v = cell_v.value
                if v is None and isinstance(cell_f.value, str):
                    m = _SUM_RE.match(cell_f.value.strip())
                    if m:
                        min_col, min_row, max_col, max_row = range_boundaries(
                            f"{m.group(1)}:{m.group(2)}"
                        )
                        total, found_any = 0, False
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                cv = ws_v.cell(row=r, column=c).value
                                if isinstance(cv, (int, float)):
                                    total += cv
                                    found_any = True
                        v = total if found_any else None
                values.append(v)
            if not any(v is not None for v in values):
                continue
            if headers is None:
                headers = [str(v) if v is not None else f"col{i}" for i, v in enumerate(values)]
                continue
            rows_out.append({headers[i]: values[i] for i in range(len(headers)) if i < len(values)})
        out[sheet_name] = rows_out
    return out


def verify_client_work_count(client_name: str) -> dict:
    """Cross-check a client's work count against a second, independent source.

    completion_certificate and company_completion_certificate are two separate
    records of the same 155 underlying works (the client's sign-off, and our own
    company record) -- verified to pair up 1:1 by package number with zero
    exceptions across the whole corpus. Their match counts for the same client
    must always be equal. Call this AFTER you believe you've found all of a
    client's works, BEFORE finalizing any count/total/list answer about them.

    If completion_certificate_count != company_completion_certificate_count,
    you have NOT found all the works yet. Compare the two doc_id lists (strip
    the "CC"/"CCC" prefix and match by number) to see which package numbers
    appear on only one side, then re-search for that missing document --
    it may use a different certificate layout than the ones you've already
    found (some certificates center/indent the client name in the letterhead
    and signature block instead of left-aligning it)."""
    cc = grep_documents(re.escape(client_name), doc_type="completion_certificate")
    ccc = grep_documents(re.escape(client_name), doc_type="company_completion_certificate")
    cc_ids = sorted(d["doc_id"] for d in cc)
    ccc_ids = sorted(d["doc_id"] for d in ccc)
    return {
        "completion_certificate_count": len(cc_ids),
        "completion_certificate_ids": cc_ids,
        "company_completion_certificate_count": len(ccc_ids),
        "company_completion_certificate_ids": ccc_ids,
        "counts_match": len(cc_ids) == len(ccc_ids),
    }


_ISO_DATE_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")


def parse_inr(val: str) -> int | None:
    """Lossless Indian-currency-rendering parser -> exact integer rupees.
    Handles 'INR 33.38 Cr', '3,338.00 Lakh', '33,38,00,000/-', '333800000'.
    Returns None if val does not parse as a recognizable money amount."""
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("/-", "").strip()
    s = re.sub(r"^(INR|Rs\.?|Rupees)\s*", "", s, flags=re.IGNORECASE).strip()
    m = re.search(r"([\d.]+)\s*(Cr|Crore)\b", s, re.IGNORECASE)
    if m:
        return round(float(m.group(1)) * 10_000_000)
    m = re.search(r"([\d.]+)\s*Lakh\b", s, re.IGNORECASE)
    if m:
        return round(float(m.group(1)) * 100_000)
    m = re.search(r"^([\d.]+)$", s)
    if m:
        return round(float(m.group(1)))
    return None


def parse_date_flexible(val: str) -> str | None:
    """Returns ISO YYYY-MM-DD or None. dayfirst=True since every OTHER date
    format observed in this corpus is Indian-convention (DD/MM/YYYY,
    "13 Nov 2020", etc.) -- ambiguous DD/MM vs MM/DD cases resolve to
    day-first.

    IMPORTANT: dates already in YYYY-MM-DD form must be returned as-is, NOT
    passed through dateutil with dayfirst=True. dateutil's dayfirst flag
    applies even to already-unambiguous ISO strings and silently swaps
    day/month whenever both are <=12 -- e.g. parse("2011-02-06",
    dayfirst=True) wrongly returns 2011-06-02 instead of 2011-02-06."""
    if not val:
        return None
    val = val.strip()
    m = _ISO_DATE_RE.match(val)
    if m:
        return val
    try:
        return dateparser.parse(val, dayfirst=True).date().isoformat()
    except (ValueError, OverflowError):
        return None


def query_db(sql: str, params: tuple = ()) -> list[dict] | dict:
    """Run a read-only SQL query against estate.db, the structured
    knowledge-layer built from this corpus's completion_certificate +
    company_completion_certificate + personnel_certificate +
    past_performance_portfolio + reference_letter documents (155 packages,
    validated 21/21 against the sample question set). Prefer this over
    grepping raw document text for anything expressible as SQL --
    aggregates, counts, thresholds, rankings, date spans across clients/
    engineers/projects -- it's far more reliable than re-deriving these by
    hand from prose every time. Still read the underlying documents
    (read_document) for fields not captured in this schema, and to verify
    a specific number before finalizing an answer.

    Schema:
      clients(client_id, canonical_name)
      engineers(engineer_id, canonical_name)
      projects(package, doc_id_cc, doc_id_ccc, client_id, project_name,
               category, value, has_discrepancy, completion_date,
               engineer_id, has_reference_letter, role)
        -- value: exact integer rupees (already parsed from Cr/Lakh/etc).
        -- completion_date: ISO 'YYYY-MM-DD', sortable as text or via
           SQLite's julianday() for day-span math.
        -- has_reference_letter: 1/0.
        -- role: 'Prime' or 'JV Partner' (from past_performance_portfolio;
           always present, 155/155).
        -- has_discrepancy: 1 if the completion_certificate and
           company_completion_certificate for this package disagree on
           value -- the row is NOT dropped, `value` picks the
           client-signed completion_certificate's number as tie-break, but
           check both doc_id_cc/doc_id_ccc directly if precision matters
           (only 1 package in the whole corpus has this flag set).
      credentials(doc_id, engineer_id, credential_id, credential_type,
                  issued_date)
        -- credential_type: 'PMP' or 'Six Sigma Black Belt'. Not every
           engineer has a credential on file (only 22/39), and some have
           both types -- filter credential_type explicitly for "PMP
           issuance date" style questions, don't assume one row per
           engineer.

    Join engineers/clients to projects via engineer_id/client_id. Client
    and engineer names may need a LIKE '%partial%' match rather than
    exact-equals -- canonical_name is the fuller/non-uppercase variant but
    questions often reference clients/engineers by a shorter form."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]
    except sqlite3.Error as e:
        return {"error": str(e)}
    finally:
        conn.close()


def find_client(name_pattern: str) -> list[dict]:
    """Shortcut for query_db("select * from clients where canonical_name
    like '%...%'") -- saves writing that boilerplate every time you need a
    client_id. Returns a list of {client_id, canonical_name} matches (0, 1,
    or several -- if more than one, pick the right one yourself, don't just
    take the first)."""
    return query_db("select * from clients where canonical_name like ?", (f"%{name_pattern}%",))


def find_engineer(name_pattern: str) -> list[dict]:
    """Same as find_client but for the engineers table."""
    return query_db("select * from engineers where canonical_name like ?", (f"%{name_pattern}%",))
