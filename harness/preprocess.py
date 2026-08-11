"""
One-time pass: extract text from every document in the corpus into a flat
text cache, so tool calls during question-answering are file reads, not
subprocess/parsing calls. Mechanical extraction only (pdftotext -layout for
PDFs, cell values for xlsx) -- no interpretation, nothing lossy.
"""
import csv
import os
import re
import subprocess
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import range_boundaries

SUM_RE = re.compile(r"^=SUM\(\s*([A-Za-z]+\d+):([A-Za-z]+\d+)\s*\)$", re.IGNORECASE)

HERE = Path(__file__).resolve().parent
# Dataset location is independent of where this harness folder sits -- set
# DATASET_ROOT to wherever you cloned BITS-Hackathon-Dataset on this machine.
# The fallback only makes sense on the machine this was first written on.
ROOT = Path(os.environ.get("DATASET_ROOT", r"C:\Users\NewGr\Downloads\BITS-Hackathon-Dataset"))
DOCS = ROOT / "documents"
INDEX_CSV = ROOT / "document_index.csv"
CACHE = HERE / "text_cache"
CACHE.mkdir(exist_ok=True)


def extract_pdf(path: Path) -> str:
    r = subprocess.run(["pdftotext", "-layout", str(path), "-"],
                        capture_output=True, text=True, encoding="utf-8", errors="replace")
    return r.stdout


def extract_xlsx(path: Path) -> str:
    """Cell values, with one fix: SUM() formula cells that have no cached
    result (common in programmatically-generated workbooks that were never
    opened in Excel to compute+cache formulas -- observed on every TOTAL row
    checked in this corpus) are recomputed directly from the same SUM range's
    values, rather than left blank. Anything not a simple single-range SUM()
    is left as-is (mechanical extraction only, nothing guessed)."""
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    out = []
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]
        out.append(f"=== SHEET: {sheet_name} ===")
        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            values = []
            for cell_v, cell_f in zip(row_v, row_f):
                v = cell_v.value
                if v is None and isinstance(cell_f.value, str):
                    m = SUM_RE.match(cell_f.value.strip())
                    if m:
                        min_col, min_row, max_col, max_row = range_boundaries(
                            f"{m.group(1)}:{m.group(2)}"
                        )
                        total = 0
                        found_any = False
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                cv = ws_v.cell(row=r, column=c).value
                                if isinstance(cv, (int, float)):
                                    total += cv
                                    found_any = True
                        v = total if found_any else None
                values.append(v)
            if any(c is not None for c in values):
                out.append("\t".join("" if c is None else str(c) for c in values))
    return "\n".join(out)


def main():
    rows = list(csv.DictReader(open(INDEX_CSV, encoding="utf-8")))
    print(f"{len(rows)} documents in index")

    done, failed = 0, []
    for row in rows:
        doc_id = row["doc_id"]
        rel_path = row["filename"]
        src = DOCS / rel_path
        out_path = CACHE / f"{doc_id}.txt"
        if out_path.exists():
            done += 1
            continue
        try:
            if src.suffix.lower() == ".pdf":
                text = extract_pdf(src)
            elif src.suffix.lower() == ".xlsx":
                text = extract_xlsx(src)
            else:
                failed.append((doc_id, "unknown extension"))
                continue
            out_path.write_text(text, encoding="utf-8")
            done += 1
        except Exception as e:
            failed.append((doc_id, str(e)))

    print(f"extracted: {done}, failed: {len(failed)}")
    for doc_id, err in failed[:20]:
        print(f"  FAILED {doc_id}: {err}")


if __name__ == "__main__":
    main()
